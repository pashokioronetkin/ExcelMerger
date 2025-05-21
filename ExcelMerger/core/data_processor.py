import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from . import constants
from .utils import letter_to_number, number_to_letter, form_date, form_date_add_id

class DataProcessor:
    def __init__(self):
        self.source_file = None
        self.target_file = None
        self.df_source = None
        self.df_target = None

    def sheet_check(self, sheet1, sheet2):
        """Проверка соответствия листов по датам."""
        try:
            for sheet_name in sheet1:
                date = []
                sheet1_day, sheet1_month = sheet_name.split()
                month_in_russian, _ = sheet2.split()

                if any(month_in_russian in s for s in constants.MONTHS):
                    sheet_data = pd.read_excel(
                        self.target_file,
                        header=None,
                        dtype=object,
                        engine='openpyxl',
                        sheet_name=sheet2
                    )

                    for _, row in sheet_data.iterrows():
                        if pd.isna(row[1]):
                            date.append(str(row[0]).split()[0])

                    self.date_check(sheet1_day, month_in_russian, form_date(date))
                else:
                    print("Месяц не найден в списке.")
        except ValueError as e:
            print(f"Ошибка формата ввода: {e}. Пример: '01 DEC'")

    def date_check(self, sheet1_day, sheet1_month, dates):
        """Проверка соответствия дат между листами."""
        for date_str in dates:
            try:
                day, month_num, year = date_str.split(".")
                month_num_from_rus = constants.MONTH_RUS_TO_NUM.get(sheet1_month, "00")

                if sheet1_day == day and month_num == month_num_from_rus:
                    print("Даты совпадают")
                else:
                    print("ОШИБКА: Даты нет в листе")
            except ValueError as e:
                print(f"Ошибка при разборе даты {date_str}: {e}")

    def add_id(self, target_sheet, id_mapping, pas_col_sheet2, id_col_sheet2):
        """Добавление ID в целевой файл."""
        try:
            wb = load_workbook(self.target_file)
            ws = wb[target_sheet]

            if isinstance(pas_col_sheet2, int):
                pas_col_index = pas_col_sheet2 - 1
            else:
                pas_col_index = letter_to_number(pas_col_sheet2) - 1

            id_col_letter = id_col_sheet2.upper() if isinstance(id_col_sheet2, str) else number_to_letter(id_col_sheet2)

            current_date = None

            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row):
                    continue

                if len(row) > 1 and (row[1] is None or pd.isna(row[1])):
                    try:
                        if hasattr(row[0], 'date'):
                            date_obj = row[0].date()
                        else:
                            date_obj = datetime.strptime(str(row[0]), "%Y-%m-%d").date()

                        formatted_date = form_date_add_id(date_obj)
                        if formatted_date:
                            day, month, _ = formatted_date.split('.')
                            en_m = constants.MONTH_NUM_TO_EN.get(month, "")
                            current_date = f"{day} {en_m}"
                    except Exception as e:
                        print(f"Не удалось распознать дату в строке {i}: {e}")
                        continue

                if current_date and current_date in id_mapping and len(row) > pas_col_index:
                    try:
                        passport = str(row[pas_col_index]).strip()
                        if passport.endswith('.0'):
                            passport = passport[:-2]

                        if passport and passport in id_mapping[current_date]:
                            ws[f'{id_col_letter}{i}'] = id_mapping[current_date][passport]
                            print(f"Добавлен ID {id_mapping[current_date][passport]} в ячейку {id_col_letter}{i}")
                    except Exception as e:
                        print(f"Ошибка обработки строки {i}: {e}")

            wb.save(f"{self.target_file.split('/')[-1]}")
            print(f"Файл успешно сохранен: {self.target_file.split('/')[-1]}")

        except Exception as e:
            print(f"Ошибка при добавлении ID: {e}")
            raise