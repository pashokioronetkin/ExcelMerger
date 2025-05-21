import pandas as pd
from PyQt6.QtWidgets import (
    QMainWindow, QPushButton, QVBoxLayout, QWidget,
    QLabel, QHBoxLayout, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from openpyxl.reader.excel import load_workbook

from core.data_processor import DataProcessor
from core.utils import letter_to_number
from ui.dialogs import choose_column_dialog, choose_sheet_dialog


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.processor = DataProcessor()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Универсальный обработчик Excel")
        self.setFixedSize(400, 300)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        self.lbl_source = QLabel("Файл с ID: не выбран")
        self.lbl_target = QLabel("Файл агента: не выбран")
        self.btn_load_source = QPushButton("Загрузить файл с ID")
        self.btn_load_target = QPushButton("Загрузить файл агента")
        self.btn_process = QPushButton("Запустить обработку")

        for label in [self.lbl_source, self.lbl_target]:
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        for btn in [self.btn_load_source, self.btn_load_target, self.btn_process]:
            btn.setFixedSize(200, 80)

        h_layouts = [QHBoxLayout() for _ in range(3)]
        for layout in h_layouts:
            layout.addStretch()

        h_layouts[0].addWidget(self.btn_load_source)
        h_layouts[1].addWidget(self.btn_load_target)
        h_layouts[2].addWidget(self.btn_process)

        for layout in h_layouts:
            layout.addStretch()

        main_layout.addWidget(self.lbl_source)
        main_layout.addLayout(h_layouts[0])
        main_layout.addWidget(self.lbl_target)
        main_layout.addLayout(h_layouts[1])
        main_layout.addLayout(h_layouts[2])
        main_layout.addStretch()

        self.btn_load_source.clicked.connect(lambda: self.load_file('source'))
        self.btn_load_target.clicked.connect(lambda: self.load_file('target'))
        self.btn_process.clicked.connect(self.process_data)

    def load_file(self, file_type):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_name:
            if file_type == 'source':
                self.processor.source_file = file_name
                self.lbl_source.setText(f"Файл с ID: {file_name.split('/')[-1]}")
            else:
                self.processor.target_file = file_name
                self.lbl_target.setText(f"Файл агента: {file_name.split('/')[-1]}")

    def process_data(self):
        if not self.processor.source_file or not self.processor.target_file:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, загрузите оба файла!")
            return

        try:
            self.processor.df_source = pd.read_excel(
                self.processor.source_file,
                header=None,
                dtype=object,
                engine='openpyxl'
            )

            day_with_id_mapping = {}

            column_source = choose_column_dialog(
                self,
                "Колонка сравнения (файл с ID)",
                "Введите колонку для сравнения (A-ZZZ):",
                "C"
            )
            if not column_source:
                return

            column_target = choose_column_dialog(
                self,
                "Колонка сравнения (файл агента)",
                "Введите колонку для сравнения (A-ZZZ):",
                "F"
            )
            if not column_target:
                return

            column_target_id = choose_column_dialog(
                self,
                "Колонка для ID",
                "Введите колонку для вставки ID (A-ZZZ):",
                "G"
            )
            if not column_target_id:
                return

            self.processor.df_target = pd.read_excel(
                self.processor.target_file,
                header=None,
                dtype=object,
                engine='openpyxl'
            )

            wb1 = load_workbook(self.processor.source_file)
            wb2 = load_workbook(self.processor.target_file)

            source_sheet = choose_sheet_dialog(
                self,
                wb1.sheetnames,
                is_multi=True
            )
            if not source_sheet:
                return

            target_sheet = choose_sheet_dialog(
                self,
                wb2.sheetnames,
                is_multi=False
            )
            if not target_sheet:
                return

            for sheet_name in source_sheet:
                id_mapping = {}
                sheet_data = pd.read_excel(
                    self.processor.source_file,
                    header=None,
                    dtype=object,
                    engine='openpyxl',
                    sheet_name=sheet_name
                )

                for _, row in sheet_data.iterrows():
                    passport = str(row[letter_to_number(column_source) - 1]).strip().replace('.0', '') if not pd.isna(
                        row[letter_to_number(column_source) - 1]) else ''
                    if passport:
                        id_mapping[passport] = row[0]
                        day_with_id_mapping[sheet_name] = id_mapping

            self.processor.sheet_check(source_sheet, target_sheet)
            self.processor.add_id(target_sheet, day_with_id_mapping, column_target, column_target_id)

            QMessageBox.information(self, "Успех", "Обработка завершена успешно!")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки:\n{str(e)}")
            print(f"Error: {e}")
