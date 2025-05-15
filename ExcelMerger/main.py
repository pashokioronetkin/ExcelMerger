import sys
import pandas as pd
import re
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,
    QFileDialog, QMessageBox, QLabel, QHBoxLayout, QInputDialog
)
from PyQt6.QtCore import Qt
from datetime import datetime

# ========== КОНСТАНТЫ И СЛОВАРИ ==========

# Кортеж с названиями месяцев на русском с годом
month = (
    "ЯНВАРЬ 2025", "ФЕВРАЛЬ 2025", "МАРТ 2025", "АПРЕЛЬ 2025", "МАЙ 2025",
    "ИЮНЬ 2025", "ИЮЛЬ 2025", "АВГУСТ 2025", "СЕНТЯБРЬ 2025", "ОКТЯБРЬ 2025",
    "НОЯБРЬ 2025", "ДЕКАБРЬ 2025")

# Словарь для перевода сокращений месяцев с английского на русский
month_en_to_rus = {
    "JAN": "ЯНВАРЬ", "FEB": "ФЕВРАЛЬ", "MAR": "МАРТ", "APR": "АПРЕЛЬ",
    "MAY": "МАЙ", "JUN": "ИЮНЬ", "JUL": "ИЮЛЬ", "AUG": "АВГУСТ",
    "SEP": "СЕНТЯБРЬ", "OCT": "ОКТЯБРЬ", "NOV": "НОЯБРЬ", "DEC": "ДЕКАБРЬ"
}

# Словарь для перевода русских названий месяцев в их числовое представление
month_rus_to_num = {
    "ЯНВАРЬ": "01", "ФЕВРАЛЬ": "02", "МАРТ": "03", "АПРЕЛЬ": "04",
    "МАЙ": "05", "ИЮНЬ": "06", "ИЮЛЬ": "07", "АВГУСТ": "08",
    "СЕНТЯБРЬ": "09", "ОКТЯБРЬ": "10", "НОЯБРЬ": "11", "ДЕКАБРЬ": "12"
}

# Словарь для перевода числового представления месяцев в английские сокращения
month_num_to_en = {
    "01": "JAN", "02": "FEB", "03": "MAR", "04": "APR",
    "05": "MAY", "06": "JUN", "07": "JUL", "08": "AUG",
    "09": "SEP", "10": "OCT", "11": "NOV", "12": "DEC"
}


class UniversalExcelUpdater(QMainWindow):
    """Главный класс приложения для обработки Excel файлов."""

    def __init__(self):
        """Инициализация главного окна приложения."""
        super().__init__()
        self.init_ui()  # Настройка интерфейса
        # Инициализация переменных для хранения путей к файлам
        self.source_file = None  # Путь к исходному файлу (с ID)
        self.target_file = None  # Путь к целевому файлу (для обновления)
        self.df_source = None  # DataFrame для исходного файла
        self.df_target = None  # DataFrame для целевого файла

    # ========== МЕТОДЫ ДЛЯ РАБОТЫ С ГРАФИЧЕСКИМ ИНТЕРФЕЙСОМ ==========

    def init_ui(self):
        """Настройка пользовательского интерфейса."""
        self.setWindowTitle("Универсальный обработчик Excel")
        self.setFixedSize(400, 300)  # Фиксированный размер окна

        # Создание центрального виджета и основного layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Создание элементов интерфейса
        self.lbl_source = QLabel("Файл EX1: не выбран")
        self.lbl_target = QLabel("Файл EX2: не выбран")
        self.btn_load_source = QPushButton("Загрузить файл с ID (EX1)")
        self.btn_load_target = QPushButton("Загрузить целевой файл (EX2)")
        self.btn_process = QPushButton("Запустить обработку")

        # Настройка стилей и выравнивания
        for label in [self.lbl_source, self.lbl_target]:
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Установка фиксированного размера для кнопок
        for btn in [self.btn_load_source, self.btn_load_target, self.btn_process]:
            btn.setFixedSize(200, 80)

        # Создание горизонтальных layout'ов для центрирования кнопок
        h_layouts = [QHBoxLayout() for _ in range(3)]
        for layout in h_layouts:
            layout.addStretch()

        # Добавление кнопок в layout'ы
        h_layouts[0].addWidget(self.btn_load_source)
        h_layouts[1].addWidget(self.btn_load_target)
        h_layouts[2].addWidget(self.btn_process)

        for layout in h_layouts:
            layout.addStretch()

        # Сборка основного интерфейса
        main_layout.addWidget(self.lbl_source)
        main_layout.addLayout(h_layouts[0])
        main_layout.addWidget(self.lbl_target)
        main_layout.addLayout(h_layouts[1])
        main_layout.addLayout(h_layouts[2])
        main_layout.addStretch()

        # Подключение сигналов кнопок
        self.btn_load_source.clicked.connect(lambda: self.load_file('source'))
        self.btn_load_target.clicked.connect(lambda: self.load_file('target'))
        self.btn_process.clicked.connect(self.process_data)

    def load_file(self, file_type):
        """
        Загрузка файла через диалоговое окно.

        Args:
            file_type (str): Тип файла ('source' или 'target')
        """
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_name:
            if file_type == 'source':
                self.source_file = file_name
                self.lbl_source.setText(f"EX1: {file_name.split('/')[-1]}")
            else:
                self.target_file = file_name
                self.lbl_target.setText(f"EX2: {file_name.split('/')[-1]}")

    # ========== ОСНОВНЫЕ МЕТОДЫ ОБРАБОТКИ ДАННЫХ ==========

    def process_data(self):
        """Основной метод обработки данных."""
        # Проверка, что оба файла загружены
        if not self.source_file or not self.target_file:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, загрузите оба файла!")
            return

        try:
            # Чтение исходного файла
            self.df_source = pd.read_excel(self.source_file, header=None, dtype=object, engine='openpyxl')
            day_with_id_mapping = {}  # Словарь для хранения соответствий ID и дат

            # Запрос колонок для обработки через диалоговые окна
            column_source = self.letter_to_number(self.choose_column_source_dialog())
            if not column_source:
                return

            column_target = self.choose_column_target_dialog()
            if not column_target:
                return

            column_target_id = self.choose_column_target_id_dialog()
            if not column_target_id:
                return

            # Чтение целевого файла
            self.df_target = pd.read_excel(self.target_file, header=None, dtype=object, engine='openpyxl')

            # Открытие файлов с помощью openpyxl для работы с листами
            wb1 = load_workbook(self.source_file)
            wb2 = load_workbook(self.target_file)

            # Выбор листов через диалоговые окна
            source_sheet = self.choose_source_sheet_dialog(wb1.sheetnames)
            if not source_sheet:
                return

            target_sheet = self.choose_target_sheet_dialog(wb2.sheetnames)
            if not target_sheet:
                return

            # Заполнение словаря сопоставлений ID и паспортных данных
            for sheet_name in source_sheet:
                id_mapping = {}
                sheet_data = pd.read_excel(
                    self.source_file,
                    header=None,
                    dtype=object,
                    engine='openpyxl',
                    sheet_name=sheet_name
                )

                for _, row in sheet_data.iterrows():
                    # Обработка паспортных данных (удаление .0 и лишних пробелов)
                    passport = str(row[column_source - 1]).strip().replace('.0', '') if not pd.isna(
                        row[column_source - 1]) else ''
                    if passport:
                        id_mapping[passport] = row[0]  # Сопоставление паспорта с ID
                        day_with_id_mapping[sheet_name] = id_mapping


            # Проверка соответствия листов
            self.sheet_check(source_sheet, target_sheet)

            # Добавление ID в целевой файл
            self.add_id(target_sheet, day_with_id_mapping, column_target, column_target_id)

            QMessageBox.information(self, "Успех", "Обработка завершена успешно!")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки:\n{str(e)}")
            print(f"Error: {e}")

    # ========== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ==========

    def letter_to_number(self, column_letter):
        """
        Конвертация буквенного обозначения колонки в числовое.

        Args:
            column_letter (str): Буквенное обозначение колонки (A, B, ..., Z, AA, AB, ...)

        Returns:
            int: Числовое представление колонки (1 для A, 2 для B, ...)
        """
        if not column_letter:
            return None

        num = 0
        for i, char in enumerate(reversed(column_letter.upper())):
            num += (ord(char) - ord('A') + 1) * (26 ** i)
        return num

    def number_to_letter(self, n):
        """
        Конвертация числового обозначения колонки в буквенное.

        Args:
            n (int): Номер колонки (1 для A, 2 для B, ...)

        Returns:
            str: Буквенное представление колонки
        """
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    # ========== МЕТОДЫ ДЛЯ РАБОТЫ С ДИАЛОГОВЫМИ ОКНАМИ ==========

    def choose_column_source_dialog(self):
        """Диалог для выбора колонки с паспортными данными в исходном файле."""
        text, ok = QInputDialog.getText(
            self,
            "Ввод названия колонки сравнения исходного файла",
            "Введите колонку для сравнения из исходного файла(A-ZZZ):"
        )
        return text if ok and text else None

    def choose_column_target_dialog(self):
        """Диалог для выбора колонки с паспортными данными в целевом файле."""
        text, ok = QInputDialog.getText(
            self,
            "Ввод названия колонки сравнения изменяемого файла",
            "Введите колонку для сравнения для изменяемого файла(A-ZZZ):"
        )
        return self.letter_to_number(text) if ok and text else None

    def choose_column_target_id_dialog(self):
        """Диалог для выбора колонки, куда вставлять ID."""
        text, ok = QInputDialog.getText(
            self,
            "Ввод названия колонки (куда вставить id)",
            "Введите колонку для вставки id)(A-ZZZ):"
        )
        return text.upper() if ok and text else None

    def choose_source_sheet_dialog(self, items):
        """Диалог для выбора листа с ID в исходном файле."""
        if not items:
            QMessageBox.warning(self, "Ошибка", "В файле нет листов!")
            return None

        item, ok = QInputDialog.getText(
            self,
            "Выбор из листа с ID",
            f"Введите названия листов через запятую (доступные: {', '.join(items)}):"
        )

        if ok and item:
            selected_sheets = [i.strip() for i in item.split(",")]
            # Проверка существования выбранных листов
            for sheet in selected_sheets:
                if sheet not in items:
                    QMessageBox.warning(self, "Ошибка", f"Лист '{sheet}' не найден в файле!")
                    return None
            return selected_sheets
        return None

    def choose_target_sheet_dialog(self, items):
        """Диалог для выбора целевого листа."""
        if not items:
            QMessageBox.warning(self, "Ошибка", "В файле нет листов!")
            return None

        item, ok = QInputDialog.getItem(
            self,
            "Выбор из списка",
            "Выберите вариант:",
            items,
            0,
            False
        )
        return item if ok and item else None

    # ========== МЕТОДЫ ДЛЯ РАБОТЫ С ДАТАМИ ==========

    def sheet_check(self, sheet1, sheet2):
        """
        Проверка соответствия листов по датам.

        Args:
            sheet1 (list): Список листов из исходного файла
            sheet2 (str): Название листа из целевого файла
        """
        try:
            for sheet_name in sheet1:
                date = []
                sheet1_day, sheet1_month = sheet_name.split()
                month_in_russian, _ = sheet2.split()

                if any(month_in_russian in s for s in month):
                    # Сбор дат из целевого файла
                    sheet_data = pd.read_excel(
                        self.target_file,
                        header=None,
                        dtype=object,
                        engine='openpyxl',
                        sheet_name=sheet2
                    )

                    for _, row in sheet_data.iterrows():
                        if pd.isna(row[1]):  # Если вторая колонка пустая - это строка с датой
                            date.append(str(row[0]).split()[0])

                    # Проверка соответствия дат
                    self.date_check(sheet1_day, month_in_russian, self.form_date(date))
                else:
                    print("Месяц не найден в списке.")
        except ValueError as e:
            print(f"Ошибка формата ввода: {e}. Пример: '01 DEC'")

    def date_check(self, sheet1_day, sheet1_month, dates):
        """
        Проверка соответствия дат между листами.

        Args:
            sheet1_day (str): День из исходного листа
            sheet1_month (str): Месяц из исходного листа
            dates (list): Список дат из целевого листа
        """
        for date_str in dates:
            try:
                day, month_num, year = date_str.split(".")
                month_num_from_rus = month_rus_to_num.get(sheet1_month, "00")

                if sheet1_day == day and month_num == month_num_from_rus:
                    print("Даты совпадают")
                else:
                    print("ОШИБКА: Даты нет в листе")
            except ValueError as e:
                print(f"Ошибка при разборе даты {date_str}: {e}")

    def form_date(self, dates):
        """
        Форматирование дат из формата YYYY-MM-DD в DD.MM.YYYY.

        Args:
            dates (list): Список дат в строковом формате

        Returns:
            list: Список отформатированных дат
        """
        date_ = []
        for i in dates:
            try:
                date_obj = datetime.strptime(i, "%Y-%m-%d")
                formatted_date = date_obj.strftime("%d.%m.%Y")
                date_.append(formatted_date)
            except ValueError as e:
                print(f"Ошибка форматирования даты {i}: {e}")
        return date_

    def form_date_add_id(self, date_str):
        """
        Форматирование одиночной даты из строки в объект datetime.

        Args:
            date_str (str): Дата в строковом формате

        Returns:
            str: Отформатированная дата или None при ошибке
        """
        try:
            date_obj = datetime.strptime(str(date_str), "%Y-%m-%d")
            return date_obj.strftime("%d.%m.%Y")
        except ValueError as e:
            print(f"Ошибка форматирования даты {date_str}: {e}")
            return None

    # ========== ОСНОВНОЙ МЕТОД ДОБАВЛЕНИЯ ID ==========

    def add_id(self, target_sheet, id_mapping, pas_col_sheet2, id_col_sheet2):
        """
        Добавление ID в целевой файл на основе сопоставлений.

        Args:
            target_sheet (str): Название целевого листа
            id_mapping (dict): Словарь сопоставлений ID
            pas_col_sheet2: Колонка с паспортными данными в целевом файле
            id_col_sheet2: Колонка для вставки ID в целевом файле
        """
        try:
            # Открытие целевого файла
            wb = load_workbook(self.target_file)
            ws = wb[target_sheet]

            # Преобразование обозначений колонок
            if isinstance(pas_col_sheet2, int):
                pas_col_index = pas_col_sheet2 - 1  # Преобразование в 0-based индекс
            else:
                pas_col_index = self.letter_to_number(pas_col_sheet2) - 1

            id_col_letter = id_col_sheet2.upper() if isinstance(id_col_sheet2, str) else self.number_to_letter(
                id_col_sheet2)

            current_date = None  # Текущая обрабатываемая дата

            # Обработка каждой строки в листе
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row):  # Пропуск пустых строк
                    continue

                # Проверка, является ли строка строкой с датой
                if len(row) > 1 and (row[1] is None or pd.isna(row[1])):
                    try:
                        # Обработка даты (как объекта datetime или строки)
                        if hasattr(row[0], 'date'):
                            date_obj = row[0].date()  # Для объектов datetime
                        else:
                            date_obj = datetime.strptime(str(row[0]), "%Y-%m-%d").date()  # Для строк

                        formatted_date = self.form_date_add_id(date_obj)
                        if formatted_date:
                            day, month, _ = formatted_date.split('.')
                            en_m = month_num_to_en.get(month, "")
                            current_date = f"{day} {en_m}"
                    except Exception as e:
                        print(f"Не удалось распознать дату в строке {i}: {e}")
                        continue

                # Обработка строк с данными
                if current_date and current_date in id_mapping and len(row) > pas_col_index:
                    try:
                        passport = str(row[pas_col_index]).strip()
                        if passport.endswith('.0'):
                            passport = passport[:-2]  # Удаление .0 в конце

                        if passport and passport in id_mapping[current_date]:
                            # Вставка ID в указанную колонку
                            ws[f'{id_col_letter}{i}'] = id_mapping[current_date][passport]
                            print(f"Добавлен ID {id_mapping[current_date][passport]} в ячейку {id_col_letter}{i}")
                    except Exception as e:
                        print(f"Ошибка обработки строки {i}: {e}")

            # Сохранение изменений
            wb.save('EX2.xlsx')
            print(f"Файл успешно сохранен: {self.target_file.split('/')[-1]}")

        except Exception as e:
            print(f"Ошибка при добавлении ID: {e}")
            raise


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UniversalExcelUpdater()
    window.show()
    sys.exit(app.exec())